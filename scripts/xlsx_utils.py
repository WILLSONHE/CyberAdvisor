"""Excel 写入与数字显示格式（千位分隔）。"""
from __future__ import annotations

import os
import sys
import tempfile
import time

from openpyxl import load_workbook

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)
from portfolio_utils import format_code_for_excel, parse_code_from_excel_cell

MONEY_FMT = "#,##0.00"
INT_FMT = "#,##0"

PORTFOLIO_MONEY_COLS = ("成本",)
PORTFOLIO_INT_COLS = ("股数",)

SIM_MONEY_COLS = ("成本", "现价", "市值", "盈亏")
SIM_INT_COLS = ("股数", "持仓时间(天)")
TEXT_COLS = ("代码",)


def _save_workbook(wb, path: str) -> None:
    """Save via temp file + replace — avoids Windows Errno 22 on in-place overwrite."""
    path = os.path.abspath(path)
    directory = os.path.dirname(path) or "."
    last_err: OSError | None = None
    try:
        for attempt in range(3):
            fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=directory)
            os.close(fd)
            try:
                wb.save(tmp)
                os.replace(tmp, path)
                return
            except OSError as e:
                last_err = e
                if os.path.exists(tmp):
                    try:
                        os.remove(tmp)
                    except OSError:
                        pass
                if attempt < 2:
                    time.sleep(0.25 * (attempt + 1))
            except Exception:
                if os.path.exists(tmp):
                    try:
                        os.remove(tmp)
                    except OSError:
                        pass
                raise
        if last_err:
            raise last_err
    finally:
        wb.close()


def apply_column_formats(
    path: str,
    *,
    money_cols: tuple[str, ...] = (),
    int_cols: tuple[str, ...] = (),
    text_cols: tuple[str, ...] = (),
    sheet: str | None = None,
) -> None:
    """为已存在的 xlsx 设置列数字格式（不改变单元格数值）。"""
    wb = load_workbook(path)
    ws = wb[sheet or wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    col_idx = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

    for col_name in money_cols:
        ci = col_idx.get(col_name)
        if not ci:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ci)
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT

    for col_name in int_cols:
        ci = col_idx.get(col_name)
        if not ci:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ci)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = INT_FMT

    for col_name in text_cols:
        ci = col_idx.get(col_name)
        if not ci:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ci)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            cell.value = format_code_for_excel(parse_code_from_excel_cell(cell.value))
            cell.number_format = "@"

    _save_workbook(wb, path)


def write_dataframe_xlsx(
    df,
    path: str,
    *,
    money_cols: tuple[str, ...] = (),
    int_cols: tuple[str, ...] = (),
    text_cols: tuple[str, ...] = TEXT_COLS,
) -> None:
    out = df.copy()
    if "代码" in out.columns:
        out["代码"] = out["代码"].apply(lambda c: format_code_for_excel(c))
    out.to_excel(path, index=False)
    apply_column_formats(path, money_cols=money_cols, int_cols=int_cols, text_cols=text_cols)


def format_portfolio_xlsx(path: str) -> None:
    apply_column_formats(path, money_cols=PORTFOLIO_MONEY_COLS, int_cols=PORTFOLIO_INT_COLS)


def format_sim_xlsx(path: str) -> None:
    apply_column_formats(
        path, money_cols=SIM_MONEY_COLS, int_cols=SIM_INT_COLS, text_cols=TEXT_COLS
    )
