"""股价预测复盘 Excel（Wiki/数据/股价预测追踪/复盘/数据复盘.xlsx）。"""
from __future__ import annotations

import os
import sys
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from outlook_deviation_reason import GAP_DECIMALS, GAP_THRESHOLD_PCT, build_deviation_reason
from outlook_paths import REVIEW_DIR, REVIEW_XLSX_COLUMNS, REVIEW_XLSX_PATH
from portfolio_utils import format_code_for_excel
from xlsx_utils import _save_workbook, MONEY_FMT

HORIZON_KEYS = ("1d", "3d", "7d")
HORIZON_LABELS = {"1d": "1日", "3d": "3日", "7d": "7日"}
HORIZON_ORDER = {"1日": 0, "3日": 1, "7日": 2}

COL = {name: i + 1 for i, name in enumerate(REVIEW_XLSX_COLUMNS)}


def _parse_date(s: Any) -> date | None:
    if s is None:
        return None
    try:
        if isinstance(s, datetime):
            return s.date()
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _ensure_workbook(*, migrate: bool = True) -> tuple[Any, Any]:
    os.makedirs(REVIEW_DIR, exist_ok=True)
    if os.path.isfile(REVIEW_XLSX_PATH):
        wb = load_workbook(REVIEW_XLSX_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if headers != list(REVIEW_XLSX_COLUMNS):
            _migrate_headers(ws)
        if migrate and ws.max_row > 1:
            _resort_worksheet(ws)
            _save_workbook(wb, REVIEW_XLSX_PATH)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, name in enumerate(REVIEW_XLSX_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
    _save_workbook(wb, REVIEW_XLSX_PATH)
    wb = load_workbook(REVIEW_XLSX_PATH)
    return wb, wb.active


def _migrate_headers(ws) -> None:
    """旧 8 列 → 新 9 列（插入「预测周期」）。"""
    old_headers = [cell.value for cell in ws[1]]
    if old_headers == list(REVIEW_XLSX_COLUMNS):
        return
    rows: list[list[Any]] = []
    has_horizon = "预测周期" in old_headers
    for r in range(2, ws.max_row + 1):
        if has_horizon:
            rows.append([ws.cell(r, c).value for c in range(1, len(REVIEW_XLSX_COLUMNS) + 1)])
        else:
            rows.append(
                [
                    ws.cell(r, 1).value,
                    ws.cell(r, 2).value,
                    ws.cell(r, 3).value,
                    ws.cell(r, 4).value,
                    "1日",
                    ws.cell(r, 5).value,
                    ws.cell(r, 6).value,
                    ws.cell(r, 7).value,
                    ws.cell(r, 8).value,
                ]
            )
    for col, name in enumerate(REVIEW_XLSX_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val)


def _sort_key(row: list[Any]) -> tuple:
    tf = _parse_date(row[COL["当前日期"] - 1]) or date.min
    code = str(row[COL["标的代码"] - 1] or "").zfill(6)
    hz = str(row[COL["预测周期"] - 1] or "1日")
    due = _parse_date(row[COL["预测日期"] - 1]) or date.min
    return (-tf.toordinal(), code, HORIZON_ORDER.get(hz, 9), -due.toordinal())


def _resort_worksheet(ws) -> None:
    if ws.max_row < 2:
        return
    rows = [[ws.cell(r, c).value for c in range(1, len(REVIEW_XLSX_COLUMNS) + 1)] for r in range(2, ws.max_row + 1)]
    rows.sort(key=_sort_key)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(rows, start=2):
        _write_row(ws, i, row)


def _row_key(code: str, track_from: str, due_date: str, horizon: str) -> tuple[str, str, str, str]:
    return (
        str(code).zfill(6),
        str(track_from)[:10],
        str(due_date)[:10],
        horizon,
    )


def _find_row(ws, key: tuple[str, str, str, str]) -> int | None:
    code, tf, due, hz = key
    for row in range(2, ws.max_row + 1):
        rc = str(ws.cell(row, COL["标的代码"]).value or "").zfill(6)
        rtf = str(ws.cell(row, COL["当前日期"]).value or "")[:10]
        rdue = str(ws.cell(row, COL["预测日期"]).value or "")[:10]
        rhz = str(ws.cell(row, COL["预测周期"]).value or "")
        if (rc, rtf, rdue, rhz) == (code, tf, due, hz):
            return row
    return None


def _pred_price_from_horizon(h: dict[str, Any]) -> float | None:
    ml = h.get("most_likely") or {}
    if ml.get("price") is not None:
        return round(float(ml["price"]), 2)
    if h.get("anchor") is not None:
        return round(float(h["anchor"]), 2)
    return None


def _gap_pct(pred: float, actual: float) -> float:
    if not pred:
        return 0.0
    return round((actual - pred) / pred * 100, GAP_DECIMALS)


def _write_row(ws, row: int, values: list[Any]) -> None:
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        if c == COL["标的代码"]:
            cell.number_format = "@"
        elif c in (COL["预测价格"], COL["实际价格"]):
            if isinstance(val, (int, float)):
                cell.number_format = MONEY_FMT
        elif c == COL["差距%"]:
            if isinstance(val, (int, float)):
                cell.number_format = "0.0000"


def _insert_sorted_row(ws, values: list[Any]) -> None:
    rows = [[ws.cell(r, c).value for c in range(1, len(REVIEW_XLSX_COLUMNS) + 1)] for r in range(2, ws.max_row + 1)]
    rows.append(values)
    rows.sort(key=_sort_key)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(rows, start=2):
        _write_row(ws, i, row)


def append_prediction_rows(
    records: list[dict[str, Any]],
    *,
    horizon_keys: tuple[str, ...] = HORIZON_KEYS,
) -> int:
    """登记后追加预测行（代码+当前日期+预测日期+周期 去重；按当前日期降序排列）。"""
    if not records:
        return 0
    wb, ws = _ensure_workbook(migrate=False)
    added = 0
    new_rows: list[list[Any]] = []
    for rec in records:
        code = str(rec.get("code", "")).zfill(6)
        name = rec.get("name") or code
        for hk in horizon_keys:
            h = (rec.get("horizons") or {}).get(hk)
            if not h:
                continue
            track_from = str(h.get("track_from") or rec.get("date") or "")[:10]
            due_date = str(h.get("due_date") or "")[:10]
            hz_label = HORIZON_LABELS.get(hk, hk)
            if not track_from or not due_date:
                continue
            key = _row_key(code, track_from, due_date, hz_label)
            if _find_row(ws, key) is not None:
                continue
            pred = _pred_price_from_horizon(h)
            if pred is None:
                continue
            new_rows.append(
                [
                    format_code_for_excel(code),
                    name,
                    track_from,
                    due_date,
                    hz_label,
                    pred,
                    None,
                    None,
                    None,
                ]
            )
            added += 1
    if new_rows:
        for row_vals in new_rows:
            _insert_sorted_row(ws, row_vals)
        _save_workbook(wb, REVIEW_XLSX_PATH)
    else:
        wb.close()
    return added


def fill_review_rows(reviews: list[dict[str, Any]]) -> int:
    """复盘完成后填写实际价格、差距%（4位小数）、详细原因。"""
    if not reviews:
        return 0
    wb, ws = _ensure_workbook(migrate=False)
    updated = 0
    for item in reviews:
        rec = item["record"]
        hk = item.get("horizon") or "1d"
        hz_label = HORIZON_LABELS.get(hk, hk)
        rev = item.get("review") or {}
        if rev.get("status") != "done":
            continue
        h = rec["horizons"][hk]
        track_from_s = str(h.get("track_from") or rec.get("date") or "")[:10]
        due_date_s = str(h.get("due_date") or "")[:10]
        track_from = _parse_date(track_from_s)
        due_date = _parse_date(due_date_s)
        code = str(rec.get("code", "")).zfill(6)
        name = str(rec.get("name") or code)
        key = _row_key(code, track_from_s, due_date_s, hz_label)
        pred = _pred_price_from_horizon(h)
        actual = float(rev.get("actual_close") or 0)
        gap = _gap_pct(pred or actual, actual) if pred else None
        reason = ""
        if gap is not None and track_from and due_date:
            reason = build_deviation_reason(
                code=code,
                name=name,
                track_from=track_from,
                due_date=due_date,
                pred_price=pred or actual,
                actual=actual,
                gap_pct=gap,
                h=h,
                rev=rev,
            )
        row = _find_row(ws, key)
        row_vals = [
            format_code_for_excel(code),
            name,
            track_from_s,
            due_date_s,
            hz_label,
            pred,
            actual,
            gap,
            reason,
        ]
        if row is None:
            _insert_sorted_row(ws, row_vals)
        else:
            _write_row(ws, row, row_vals)
            if gap is not None and abs(gap) < GAP_THRESHOLD_PCT:
                ws.cell(row=row, column=COL["原因"], value="")
        updated += 1
    if updated:
        _resort_worksheet(ws)
        _save_workbook(wb, REVIEW_XLSX_PATH)
    else:
        wb.close()
    return updated


def rebuild_from_log(
    *,
    pool_only: bool = True,
    universes: tuple[str, ...] = ("track", "portfolio", "queried"),
) -> dict[str, int]:
    """从预测登记.json 重建 xlsx（迁移列、补全 1/3/7 行、排序）。"""
    from outlook_tracker import HORIZON_KEYS, _load_log
    from outlook_universe import iter_universe

    pool_codes: set[str] = set()
    if pool_only:
        for uni in universes:
            pool_codes |= {e.code for e in iter_universe(uni)}

    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(REVIEW_XLSX_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)

    recs = _load_log().get("records") or []
    if pool_only:
        recs = [r for r in recs if str(r.get("code", "")).zfill(6) in pool_codes]

    rows: list[list[Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for rec in sorted(recs, key=lambda r: r.get("date", "")):
        code = str(rec.get("code", "")).zfill(6)
        name = rec.get("name") or code
        for hk in HORIZON_KEYS:
            h = (rec.get("horizons") or {}).get(hk)
            if not h:
                continue
            tf = str(h.get("track_from") or rec.get("date") or "")[:10]
            due = str(h.get("due_date") or "")[:10]
            hz = HORIZON_LABELS.get(hk, hk)
            key = _row_key(code, tf, due, hz)
            if key in seen:
                continue
            seen.add(key)
            pred = _pred_price_from_horizon(h)
            if pred is None:
                continue
            rev = h.get("review") or {}
            actual = rev.get("actual_close") if rev.get("status") == "done" else None
            gap = _gap_pct(pred, float(actual)) if actual is not None else None
            reason = ""
            if gap is not None and actual is not None:
                tfd = _parse_date(tf)
                dud = _parse_date(due)
                if tfd and dud:
                    reason = build_deviation_reason(
                        code=code,
                        name=name,
                        track_from=tfd,
                        due_date=dud,
                        pred_price=pred,
                        actual=float(actual),
                        gap_pct=gap,
                        h=h,
                        rev=rev,
                    )
            rows.append(
                [
                    format_code_for_excel(code),
                    name,
                    tf,
                    due,
                    hz,
                    pred,
                    float(actual) if actual is not None else None,
                    gap,
                    reason,
                ]
            )

    rows.sort(key=_sort_key)
    for i, row in enumerate(rows, start=2):
        _write_row(ws, i, row)
    _save_workbook(wb, REVIEW_XLSX_PATH)
    return {"rows": len(rows), "path": REVIEW_XLSX_PATH}


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="数据复盘.xlsx 维护")
    ap.add_argument("--rebuild", action="store_true", help="从预测登记.json 重建并排序")
    ap.add_argument("--migrate", action="store_true", help="仅迁移列并排序现有文件")
    args = ap.parse_args()
    if args.rebuild:
        stat = rebuild_from_log()
        print(f"Rebuilt {stat['rows']} rows → {stat['path']}")
    elif args.migrate:
        wb, ws = _ensure_workbook(migrate=True)
        wb.close()
        print(f"Migrated → {REVIEW_XLSX_PATH}")
    else:
        from outlook_tracker import _load_log

        recs = _load_log().get("records") or []
        latest = [r for r in recs if r.get("date") == str(date.today())]
        n = append_prediction_rows(latest)
        print(f"Appended {n} rows → {REVIEW_XLSX_PATH}")
